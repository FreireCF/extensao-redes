# -*- coding: utf-8 -*-
"""
analise_internet.py
====================
Pipeline reproduzível para limpeza, padronização e análise estatística de
respostas de um formulário (Google Forms) sobre qualidade de internet em
Caxias-MA.

Basta trocar o caminho do arquivo de entrada (INPUT_FILE) para rodar a
análise novamente com uma nova exportação do Google Forms.

Saídas:
  - dados_limpos.xlsx       -> dados limpos e padronizados (sem colunas extras)
  - dados_analisados.xlsx   -> dados limpos + colunas calculadas + classificação + abas de estatísticas
  - graficos/*.png          -> gráficos em alta resolução
  - relatorio.md            -> relatório técnico completo

Autor: pipeline gerado por Claude (Anthropic)
"""

import re
import unicodedata
import warnings
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")

# ----------------------------------------------------------------------------
# CONFIGURAÇÃO GERAL
# ----------------------------------------------------------------------------

BASE_DIR = Path(__file__).resolve().parent

OUTPUT_DIR = BASE_DIR / "saida"
GRAPH_DIR = OUTPUT_DIR / "graficos"
CLEAN_DATA_FILE = OUTPUT_DIR / "dados_limpos.xlsx"
ANALYSIS_FILE = OUTPUT_DIR / "dados_analisados.xlsx"
REPORT_FILE = OUTPUT_DIR / "relatorio.md"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
GRAPH_DIR.mkdir(parents=True, exist_ok=True)


def _detectar_arquivo_entrada() -> Path:
    """Procura automaticamente o arquivo de respostas exportado do Google Forms.

    Prioridade: primeiro .xlsx encontrado; se não houver, primeiro .csv.
    Lança FileNotFoundError com mensagem amigável se nenhum for encontrado.
    """
    xlsx_files = sorted(BASE_DIR.glob("*.xlsx"))
    if xlsx_files:
        return xlsx_files[0]
    csv_files = sorted(BASE_DIR.glob("*.csv"))
    if csv_files:
        return csv_files[0]
    raise FileNotFoundError(
        f"Nenhum arquivo de respostas (.xlsx ou .csv) encontrado em:\n  {BASE_DIR}\n"
        "Exporte o formulário do Google Forms e coloque o arquivo na mesma pasta "
        "que este script."
    )


INPUT_FILE = _detectar_arquivo_entrada()

plt.rcParams.update({
    "figure.dpi": 150,
    "savefig.dpi": 300,
    "font.size": 11,
    "axes.titlesize": 13,
    "axes.titleweight": "bold",
    "axes.spines.top": False,
    "axes.spines.right": False,
})
COLOR_MAIN = "#2563eb"
COLOR_PALETTE = ["#2563eb", "#16a34a", "#f59e0b", "#dc2626", "#7c3aed",
                  "#0891b2", "#db2777", "#65a30d", "#ea580c", "#475569"]

# Mapeamento de nomes de colunas originais (Google Forms) -> nomes curtos
COLUMN_MAP = {
    "Carimbo de data/hora": "timestamp",
    "Você leu, compreendeu e concorda em participar desta pesquisa?": "consentimento",
    "Qual é o seu bairro de residência?": "bairro",
    "Qual é o seu provedor de internet?": "provedor",
    "Qual é o tipo de conexão de internet que você utiliza atualmente em sua residência?": "tipo_conexao",
    " Qual é a velocidade de internet que você contratou no seu provedor?  ": "plano_contratado_raw",
    "Qual foi a velocidade média de download da sua internet?\n(Em Mbps)": "download_m1_raw",
    "Qual foi a velocidade média de upload da sua internet?\n(Em Mbps)": "upload_m1_raw",
    "Qual foi o seu ping médio (latência)?\n(em milissengundos - ms)": "ping_m1_raw",
    "Qual foi a perda de pacotes (packet loss) observada?\n(Em %)": "perda_m1_raw",
    "Qual foi a velocidade média de download da sua internet?_x000a_(Em Mbps) 2": "download_m2_raw",
    "Qual foi a velocidade média de upload da sua internet?_x000a_(Em Mbps) 2": "upload_m2_raw",
    "Qual foi o seu ping médio (latência)?_x000a_(em milissengundos - ms) 2": "ping_m2_raw",
    "Qual foi a perda de pacotes (packet loss) observada?_x000a_(Em %) 2": "perda_m2_raw",
}


# ----------------------------------------------------------------------------
# ETAPA 1 — INSPEÇÃO
# ----------------------------------------------------------------------------
def inspecionar(df: pd.DataFrame) -> None:
    print("=" * 70)
    print("ETAPA 1 — INSPEÇÃO DA PLANILHA")
    print("=" * 70)
    print(f"Linhas: {df.shape[0]} | Colunas: {df.shape[1]}")
    vazias = [c for c in df.columns if df[c].isna().all()]
    print(f"Colunas totalmente vazias: {vazias or 'nenhuma'}")
    print("Colunas:")
    for c in df.columns:
        print(f"  - {c}")
    print()


# ----------------------------------------------------------------------------
# FUNÇÕES AUXILIARES DE LIMPEZA DE TEXTO
# ----------------------------------------------------------------------------
def strip_invisible(texto: str) -> str:
    """Remove espaços extras, caracteres invisíveis e normaliza espaços."""
    if pd.isna(texto):
        return texto
    texto = str(texto)
    # remove caracteres de controle/invisíveis (exceto espaço comum)
    texto = "".join(ch for ch in texto if unicodedata.category(ch)[0] != "C")
    # normaliza múltiplos espaços
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def normaliza_chave(texto: str) -> str:
    """Gera uma chave normalizada (sem acento, minúscula, sem espaços extras)
    para permitir comparação/padronização robusta de variações de digitação."""
    if pd.isna(texto):
        return ""
    t = strip_invisible(texto).lower()
    t = unicodedata.normalize("NFKD", t).encode("ascii", "ignore").decode("ascii")
    # remove prefixos comuns repetidamente (ex.: "residencial vila paraiso" -> "paraiso")
    prefixo = re.compile(r"^(bairro|residencial|vila do|vila|vl)\s+")
    while prefixo.match(t):
        t = prefixo.sub("", t)
    t = re.sub(r"[^a-z0-9 ]", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


# Dicionário de padronização de BAIRRO: chave normalizada -> nome final (Title Case)
BAIRRO_PADRAO = {
    "centro": "Centro",
    "ponte": "Ponte",
    "teso duro": "Teso Duro",
    "campo de belem": "Campo de Belém",
    "cangalheiro": "Cangalheiro",
    "paraiso": "Vila Paraíso",
    "caldeiroes": "Caldeirões",
    "cohab": "COHAB",
    "refinaria": "Refinaria",
    "volta redonda": "Volta Redonda",
    "seriema": "Seriema",
    "eugenio coutinho": "Residencial Eugênio Coutinho",
    "sao francisco": "São Francisco",
    "tamarineiro": "Tamarineiro",
    "tamarineiro novo": "Tamarineiro Novo",
    "castelo branco": "Castelo Branco",
    "piquizeiro": "Piquizeiro",
}


def padroniza_bairro(valor: str) -> str:
    chave = normaliza_chave(valor)
    return BAIRRO_PADRAO.get(chave, strip_invisible(valor).title() if pd.notna(valor) else valor)


def padroniza_provedor(valor: str) -> str:
    """Padroniza nomes de provedores: limpa espaços/maiúsculas mas preserva a
    grafia oficial de marca (Title Case com exceções conhecidas)."""
    if pd.isna(valor):
        return valor
    v = strip_invisible(valor)
    chave = normaliza_chave(v).replace(" ", "")
    excecoes = {
        "ibl net": "IBL NET", "iblnet": "IBL NET",
        "iprontomax": "iPronto Max",
        "netponte": "NetPonte",
        "provedorturbonet": "Provedor Turbonet",
        "turbonetcaxias": "TurboNet Caxias",
        "oknetinfor": "Ok Net Infor",
        "speednet": "SpeedNet",
        "paraisointernet": "Paraíso Internet",
        "cohabnet": "CohabNet",
        "portalmailinternet": "PortalMail Internet",
        "portalmailvilaparaiso": "Portalmail Vila Paraíso",
        "campunetcaxiasprovedordeinternet": "CampuNet Caxias - Provedor de Internet",
        "atualinternetbandalarga": "Atual Internet Banda Larga",
        "connectmaranhao": "Connect Maranhão",
        "claronet": "Claro Net",
    }
    chave_simples = normaliza_chave(v).replace(" ", "")
    if chave_simples in excecoes:
        return excecoes[chave_simples]
    return v  # mantém original se não reconhecido (preserva nomes próprios)


def padroniza_conexao(valor: str) -> str:
    if pd.isna(valor):
        return valor
    v = normaliza_chave(valor)
    if "fibra" in v and ("movel" in v or "4g" in v or "5g" in v):
        return "Fibra Óptica + Internet Móvel"
    if "fibra" in v:
        return "Fibra Óptica"
    if "cabo" in v or "coaxial" in v:
        return "Cabo (Coaxial)"
    if "nao sei" in v:
        return "Não sei informar"
    return strip_invisible(valor)


# ----------------------------------------------------------------------------
# CONVERSÃO DE FAIXAS (TEXTO) PARA VALORES NUMÉRICOS (PONTO MÉDIO)
# ----------------------------------------------------------------------------
def faixa_para_numero(valor: str, tipo: str):
    """Converte respostas em formato de faixa textual do Google Forms em um
    valor numérico representativo (ponto médio da faixa).

    tipo: 'velocidade', 'ping' ou 'perda' — define como tratar faixas abertas.
    """
    if pd.isna(valor):
        return np.nan
    v = normaliza_chave(valor)
    if v in ("nao sei informar", ""):
        return np.nan

    # caso já seja número (perda de pacotes "0" vem como int/float do Excel)
    try:
        return float(str(valor).replace(",", "."))
    except (ValueError, TypeError):
        pass

    # extrai todos os números do texto (vírgula decimal -> ponto)
    numeros = re.findall(r"\d+(?:[.,]\d+)?", v.replace(",", "."))
    numeros = [float(n) for n in numeros]

    if "mais" in v or "acima" in v or "ou mais" in v:
        # faixa aberta para cima: usa o próprio valor mínimo como referência
        return numeros[0] if numeros else np.nan
    if "menos" in v or "ate" in v:
        # faixa aberta para baixo: usa metade do valor citado como estimativa central
        return numeros[0] / 2 if numeros else np.nan
    if len(numeros) >= 2:
        return (numeros[0] + numeros[1]) / 2
    if len(numeros) == 1:
        return numeros[0]
    return np.nan


def periodo_do_dia(hora: int) -> str:
    if 5 <= hora < 12:
        return "Manhã"
    if 12 <= hora < 18:
        return "Tarde"
    return "Noite"


# ----------------------------------------------------------------------------
# ETAPA 2 — LIMPEZA
# ----------------------------------------------------------------------------
def limpar_dados(df_raw: pd.DataFrame) -> pd.DataFrame:
    print("=" * 70)
    print("ETAPA 2 — LIMPEZA DOS DADOS")
    print("=" * 70)

    df = df_raw.rename(columns=COLUMN_MAP).copy()

    n_inicial = len(df)

    # garante que timestamp seja datetime (pode vir como string dependendo do Excel)
    if "timestamp" in df.columns:
        df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")

    # remove linhas completamente vazias
    df = df.dropna(how="all")

    # remove duplicadas exatas (resposta idêntica em todas as colunas)
    n_antes = len(df)
    df = df.drop_duplicates()
    print(f"Linhas totalmente duplicadas removidas: {n_antes - len(df)}")

    # limpa espaços/caracteres invisíveis em colunas de texto (apenas as que existirem)
    colunas_texto = ["bairro", "provedor", "tipo_conexao"]
    for c in colunas_texto:
        if c not in df.columns:
            print(f"  Aviso: coluna esperada '{c}' não encontrada após renomear — verifique COLUMN_MAP.")
            continue
        df[c] = df[c].apply(strip_invisible)

    # remove coluna de consentimento (irrelevante para análise - todos concordaram)
    if "consentimento" in df.columns:
        df = df.drop(columns=["consentimento"])

    # padronização de texto
    df["bairro"] = df["bairro"].apply(padroniza_bairro)
    df["provedor"] = df["provedor"].apply(padroniza_provedor)
    df["tipo_conexao"] = df["tipo_conexao"].apply(padroniza_conexao)

    # plano contratado -> numérico (Mbps)
    df["plano_contratado_mbps"] = df["plano_contratado_raw"].apply(
        lambda x: faixa_para_numero(x, "velocidade"))

    # converte as duas medições (M1 e M2) de cada métrica para número
    for metrica, tipo in [("download", "velocidade"), ("upload", "velocidade"),
                           ("ping", "ping"), ("perda", "perda")]:
        df[f"{metrica}_m1"] = df[f"{metrica}_m1_raw"].apply(lambda x: faixa_para_numero(x, tipo))
        df[f"{metrica}_m2"] = df[f"{metrica}_m2_raw"].apply(lambda x: faixa_para_numero(x, tipo))
        # média das duas medições = valor final usado na análise
        df[f"{metrica}_mbps" if metrica in ("download", "upload") else metrica] = (
            df[[f"{metrica}_m1", f"{metrica}_m2"]].mean(axis=1, skipna=True)
        )

    df = df.rename(columns={"ping": "ping_ms", "perda": "perda_pct"})

    # horário/período da medição, a partir do timestamp do envio do formulário
    df["hora_resposta"] = df["timestamp"].dt.hour
    df["periodo"] = df["hora_resposta"].apply(periodo_do_dia)
    df["data_resposta"] = df["timestamp"].dt.date

    # remove registros sem nenhuma métrica numérica válida (impossibilita análise)
    metricas_chave = ["download_mbps", "upload_mbps", "ping_ms", "perda_pct"]
    n_antes = len(df)
    df = df.dropna(subset=metricas_chave, how="all")
    print(f"Linhas sem nenhuma métrica numérica válida removidas: {n_antes - len(df)}")

    # remove duplicatas "lógicas": mesmo bairro + provedor + timestamp idêntico
    n_antes = len(df)
    df = df.drop_duplicates(subset=["timestamp", "bairro", "provedor"])
    print(f"Duplicatas lógicas (mesmo timestamp/bairro/provedor) removidas: {n_antes - len(df)}")

    # seleciona e organiza colunas finais "limpas" (sem as colunas brutas/intermediárias)
    colunas_finais = [
        "timestamp", "data_resposta", "hora_resposta", "periodo",
        "bairro", "provedor", "tipo_conexao",
        "plano_contratado_mbps", "download_mbps", "upload_mbps",
        "ping_ms", "perda_pct",
    ]
    df_limpo = df[colunas_finais].reset_index(drop=True)

    print(f"Total de respostas: {n_inicial} -> {len(df_limpo)} após limpeza")
    print(f"Bairros únicos: {df_limpo['bairro'].nunique()}")
    print(f"Provedores únicos: {df_limpo['provedor'].nunique()}")
    print()
    return df_limpo


# ----------------------------------------------------------------------------
# ETAPA 3 — VALIDAÇÃO / DADOS SUSPEITOS
# ----------------------------------------------------------------------------
def detectar_suspeitos(df: pd.DataFrame) -> pd.DataFrame:
    print("=" * 70)
    print("ETAPA 3 — VALIDAÇÃO (DADOS SUSPEITOS)")
    print("=" * 70)

    registros = []

    def add(idx, motivo):
        row = df.loc[idx]
        registros.append({
            "indice": idx,
            "bairro": row["bairro"],
            "provedor": row["provedor"],
            "motivo": motivo,
        })

    for idx, row in df.iterrows():
        if pd.notna(row["download_mbps"]) and row["download_mbps"] < 0:
            add(idx, "Download negativo")
        if pd.notna(row["upload_mbps"]) and row["upload_mbps"] < 0:
            add(idx, "Upload negativo")
        if pd.notna(row["ping_ms"]) and row["ping_ms"] < 0:
            add(idx, "Ping negativo")
        if pd.notna(row["perda_pct"]) and (row["perda_pct"] > 100 or row["perda_pct"] < 0):
            add(idx, "Perda de pacotes fora do intervalo 0-100%")
        if pd.notna(row["download_mbps"]) and row["download_mbps"] > 1000:
            add(idx, "Velocidade de download extremamente alta (>1000 Mbps)")
        if pd.notna(row["plano_contratado_mbps"]) and pd.notna(row["download_mbps"]):
            if row["download_mbps"] > row["plano_contratado_mbps"] * 1.2:
                add(idx, "Download medido muito superior ao plano contratado (>120%)")
        if pd.notna(row["plano_contratado_mbps"]) and row["plano_contratado_mbps"] == 0:
            add(idx, "Velocidade contratada igual a zero")

    df_susp = pd.DataFrame(registros)
    print(f"Total de registros suspeitos identificados: {df_susp['indice'].nunique() if len(df_susp) else 0}")
    if len(df_susp):
        print(df_susp.to_string(index=False))
    else:
        print("Nenhum dado suspeito encontrado.")
    print()
    return df_susp


# ----------------------------------------------------------------------------
# ETAPA 4 — NOVAS COLUNAS
# ----------------------------------------------------------------------------
def criar_colunas(df: pd.DataFrame) -> pd.DataFrame:
    print("=" * 70)
    print("ETAPA 4 — CRIAÇÃO DE NOVAS COLUNAS")
    print("=" * 70)
    df = df.copy()

    df["aproveitamento_download_pct"] = np.where(
        df["plano_contratado_mbps"].notna() & (df["plano_contratado_mbps"] > 0),
        (df["download_mbps"] / df["plano_contratado_mbps"] * 100).round(1),
        np.nan,
    )
    df["aproveitamento_upload_pct"] = np.where(
        df["plano_contratado_mbps"].notna() & (df["plano_contratado_mbps"] > 0),
        (df["upload_mbps"] / df["plano_contratado_mbps"] * 100).round(1),
        np.nan,
    )
    df["ping_medio_ms"] = df["ping_ms"].round(1)
    df["categoria_conexao"] = df["tipo_conexao"]

    print("Colunas criadas: aproveitamento_download_pct, aproveitamento_upload_pct, "
          "ping_medio_ms, categoria_conexao")
    print()
    return df


# ----------------------------------------------------------------------------
# ETAPA 5 — CLASSIFICAÇÃO
# ----------------------------------------------------------------------------
def classificar(row) -> str:
    """Classifica a conexão com base no aproveitamento do plano contratado,
    no ping e na perda de pacotes, seguindo os critérios sugeridos no
    enunciado do projeto."""
    aproveitamento = row["aproveitamento_download_pct"]
    ping = row["ping_ms"]
    perda = row["perda_pct"]

    # se faltam dados essenciais, não é possível classificar
    if pd.isna(aproveitamento) or pd.isna(ping) or pd.isna(perda):
        return "Não classificado"

    if aproveitamento >= 90 and ping <= 20 and perda == 0:
        return "Ótima"
    if aproveitamento >= 75 and ping <= 40 and perda <= 1:
        return "Boa"
    if aproveitamento >= 50 and ping <= 70 and perda <= 2:
        return "Regular"
    if aproveitamento >= 30 and ping <= 100 and perda <= 5:
        return "Ruim"
    return "Péssima"


def classificar_conexoes(df: pd.DataFrame) -> pd.DataFrame:
    print("=" * 70)
    print("ETAPA 5 — CLASSIFICAÇÃO DAS CONEXÕES")
    print("=" * 70)
    df = df.copy()
    df["classificacao"] = df.apply(classificar, axis=1)
    print(df["classificacao"].value_counts())
    print()
    return df


# ----------------------------------------------------------------------------
# ETAPA 6 — ESTATÍSTICAS
# ----------------------------------------------------------------------------
ORDEM_CLASSIFICACAO = {"Ótima": 5, "Boa": 4, "Regular": 3, "Ruim": 2, "Péssima": 1, "Não classificado": 0}


def estatisticas_gerais(df: pd.DataFrame) -> pd.Series:
    geral = pd.Series({
        "qtd_respostas": len(df),
        "qtd_bairros": df["bairro"].nunique(),
        "qtd_provedores": df["provedor"].nunique(),
        "download_medio_mbps": df["download_mbps"].mean(),
        "upload_medio_mbps": df["upload_mbps"].mean(),
        "ping_medio_ms": df["ping_ms"].mean(),
        "perda_media_pct": df["perda_pct"].mean(),
    }).round(2)
    return geral


def estatisticas_por_provedor(df: pd.DataFrame) -> pd.DataFrame:
    g = df.groupby("provedor").agg(
        participantes=("provedor", "count"),
        download_medio=("download_mbps", "mean"),
        upload_medio=("upload_mbps", "mean"),
        ping_medio=("ping_ms", "mean"),
        perda_media=("perda_pct", "mean"),
        aproveitamento_medio_pct=("aproveitamento_download_pct", "mean"),
    ).round(2)
    g["classificacao_media_score"] = df.groupby("provedor")["classificacao"].apply(
        lambda s: np.mean([ORDEM_CLASSIFICACAO.get(v, 0) for v in s])).round(2)
    g = g.sort_values("classificacao_media_score", ascending=False)
    return g.reset_index()


def estatisticas_por_bairro(df: pd.DataFrame) -> pd.DataFrame:
    g = df.groupby("bairro").agg(
        participantes=("bairro", "count"),
        download_medio=("download_mbps", "mean"),
        upload_medio=("upload_mbps", "mean"),
        ping_medio=("ping_ms", "mean"),
        perda_media=("perda_pct", "mean"),
    ).round(2)
    g = g.sort_values("download_medio", ascending=False)
    return g.reset_index()


def estatisticas_por_periodo(df: pd.DataFrame) -> pd.DataFrame:
    g = df.groupby("periodo").agg(
        participantes=("periodo", "count"),
        download_medio=("download_mbps", "mean"),
        upload_medio=("upload_mbps", "mean"),
        ping_medio=("ping_ms", "mean"),
        perda_media=("perda_pct", "mean"),
    ).round(2)
    ordem = ["Manhã", "Tarde", "Noite"]
    g = g.reindex([p for p in ordem if p in g.index])
    return g.reset_index()


# ----------------------------------------------------------------------------
# ETAPA 7 — PADRÕES / INSIGHTS
# ----------------------------------------------------------------------------
def descobrir_padroes(df: pd.DataFrame, por_provedor: pd.DataFrame, por_bairro: pd.DataFrame,
                       por_periodo: pd.DataFrame) -> dict:
    insights = {}
    if len(por_provedor):
        insights["melhor_provedor"] = por_provedor.iloc[0]["provedor"]
        insights["pior_provedor"] = por_provedor.iloc[-1]["provedor"]
        insights["menor_ping_provedor"] = por_provedor.loc[por_provedor["ping_medio"].idxmin(), "provedor"]
        insights["maior_aproveitamento_provedor"] = por_provedor.loc[
            por_provedor["aproveitamento_medio_pct"].idxmax(), "provedor"]
        # maior variação (desvio padrão do download) entre provedores com >=2 respostas
        var_provedor = df.groupby("provedor")["download_mbps"].agg(["std", "count"])
        var_provedor = var_provedor[var_provedor["count"] >= 2].dropna()
        if len(var_provedor):
            insights["maior_variacao_provedor"] = var_provedor["std"].idxmax()

    if len(por_bairro):
        insights["melhor_bairro"] = por_bairro.iloc[0]["bairro"]
        insights["pior_bairro"] = por_bairro.iloc[-1]["bairro"]
        media_geral_download = df["download_mbps"].mean()
        bairros_baixo = por_bairro[por_bairro["download_medio"] < media_geral_download * 0.5]
        insights["bairros_desempenho_muito_inferior"] = bairros_baixo["bairro"].tolist()

    if len(por_periodo) > 1:
        pior_periodo = por_periodo.loc[por_periodo["ping_medio"].idxmax(), "periodo"]
        melhor_periodo = por_periodo.loc[por_periodo["ping_medio"].idxmin(), "periodo"]
        insights["periodo_maior_ping"] = pior_periodo
        insights["periodo_menor_ping"] = melhor_periodo
        diff_ping = por_periodo["ping_medio"].max() - por_periodo["ping_medio"].min()
        insights["degradacao_horario_pico"] = bool(diff_ping >= 10)  # heurística: >=10ms de diferença
        insights["diferenca_ping_periodos_ms"] = round(diff_ping, 2)

    # diferença significativa entre bairros (heurística baseada em desvio padrão das médias)
    if len(por_bairro) > 1:
        cv = por_bairro["download_medio"].std() / por_bairro["download_medio"].mean()
        insights["diferenca_significativa_bairros"] = bool(cv > 0.3)
        insights["coef_variacao_bairros"] = round(cv, 2)

    return insights


# ----------------------------------------------------------------------------
# ETAPA 8 — GRÁFICOS
# ----------------------------------------------------------------------------
def salvar(fig, nome):
    fig.tight_layout()
    fig.savefig(GRAPH_DIR / nome, dpi=300, bbox_inches="tight")
    plt.close(fig)


def grafico_barras_por_provedor(por_provedor, coluna, titulo, ylabel, nome_arquivo, cor=COLOR_MAIN):
    fig, ax = plt.subplots(figsize=(10, 6))
    dados = por_provedor.sort_values(coluna, ascending=False)
    ax.bar(dados["provedor"], dados[coluna], color=cor)
    ax.set_title(titulo)
    ax.set_ylabel(ylabel)
    ax.set_xlabel("Provedor")
    plt.xticks(rotation=45, ha="right")
    for i, v in enumerate(dados[coluna]):
        ax.text(i, v, f"{v:.1f}", ha="center", va="bottom", fontsize=9)
    salvar(fig, nome_arquivo)


def gerar_graficos(df: pd.DataFrame, por_provedor: pd.DataFrame, por_bairro: pd.DataFrame) -> None:
    print("=" * 70)
    print("ETAPA 8 — GERAÇÃO DE GRÁFICOS")
    print("=" * 70)

    grafico_barras_por_provedor(por_provedor, "download_medio", "Download Médio por Provedor",
                                 "Mbps", "01_download_medio_provedor.png", COLOR_MAIN)
    grafico_barras_por_provedor(por_provedor, "upload_medio", "Upload Médio por Provedor",
                                 "Mbps", "02_upload_medio_provedor.png", "#16a34a")
    grafico_barras_por_provedor(por_provedor, "ping_medio", "Ping Médio por Provedor",
                                 "ms", "03_ping_medio_provedor.png", "#f59e0b")
    grafico_barras_por_provedor(por_provedor, "perda_media", "Perda de Pacotes Média por Provedor",
                                 "%", "04_perda_media_provedor.png", "#dc2626")

    # distribuição de provedores (pizza/barras)
    fig, ax = plt.subplots(figsize=(8, 8))
    contagem = df["provedor"].value_counts()
    ax.pie(contagem, labels=contagem.index, autopct="%1.0f%%", colors=COLOR_PALETTE * 3,
           textprops={"fontsize": 8})
    ax.set_title("Distribuição dos Provedores entre os Respondentes")
    salvar(fig, "05_distribuicao_provedores.png")

    # participantes por bairro
    fig, ax = plt.subplots(figsize=(10, 6))
    dados = por_bairro.sort_values("participantes", ascending=False)
    ax.bar(dados["bairro"], dados["participantes"], color="#7c3aed")
    ax.set_title("Participantes por Bairro")
    ax.set_ylabel("Quantidade de respostas")
    plt.xticks(rotation=60, ha="right")
    salvar(fig, "06_participantes_por_bairro.png")

    # classificação das conexões
    fig, ax = plt.subplots(figsize=(8, 6))
    ordem = ["Ótima", "Boa", "Regular", "Ruim", "Péssima", "Não classificado"]
    contagem = df["classificacao"].value_counts().reindex(ordem).dropna()
    cores_classif = {"Ótima": "#16a34a", "Boa": "#65a30d", "Regular": "#f59e0b",
                      "Ruim": "#ea580c", "Péssima": "#dc2626", "Não classificado": "#94a3b8"}
    ax.bar(contagem.index, contagem.values, color=[cores_classif[c] for c in contagem.index])
    ax.set_title("Classificação Geral das Conexões")
    ax.set_ylabel("Quantidade")
    salvar(fig, "07_classificacao_conexoes.png")

    # boxplot download
    fig, ax = plt.subplots(figsize=(6, 6))
    ax.boxplot(df["download_mbps"].dropna(), vert=True, patch_artist=True,
               boxprops=dict(facecolor=COLOR_MAIN, alpha=0.6))
    ax.set_title("Boxplot — Velocidade de Download")
    ax.set_ylabel("Mbps")
    salvar(fig, "08_boxplot_download.png")

    # boxplot ping
    fig, ax = plt.subplots(figsize=(6, 6))
    ax.boxplot(df["ping_ms"].dropna(), vert=True, patch_artist=True,
               boxprops=dict(facecolor="#f59e0b", alpha=0.6))
    ax.set_title("Boxplot — Ping (Latência)")
    ax.set_ylabel("ms")
    salvar(fig, "09_boxplot_ping.png")

    # histograma download
    fig, ax = plt.subplots(figsize=(8, 6))
    ax.hist(df["download_mbps"].dropna(), bins=10, color=COLOR_MAIN, edgecolor="white")
    ax.set_title("Histograma — Velocidade de Download")
    ax.set_xlabel("Mbps")
    ax.set_ylabel("Frequência")
    salvar(fig, "10_histograma_download.png")

    # histograma ping
    fig, ax = plt.subplots(figsize=(8, 6))
    ax.hist(df["ping_ms"].dropna(), bins=10, color="#f59e0b", edgecolor="white")
    ax.set_title("Histograma — Ping (Latência)")
    ax.set_xlabel("ms")
    ax.set_ylabel("Frequência")
    salvar(fig, "11_histograma_ping.png")

    # heatmap de correlação
    colunas_num = ["plano_contratado_mbps", "download_mbps", "upload_mbps", "ping_ms",
                   "perda_pct", "aproveitamento_download_pct"]
    corr = df[colunas_num].corr()
    fig, ax = plt.subplots(figsize=(8, 7))
    im = ax.imshow(corr, cmap="RdBu_r", vmin=-1, vmax=1)
    ax.set_xticks(range(len(corr.columns)))
    ax.set_yticks(range(len(corr.columns)))
    ax.set_xticklabels(corr.columns, rotation=45, ha="right")
    ax.set_yticklabels(corr.columns)
    for i in range(len(corr.columns)):
        for j in range(len(corr.columns)):
            ax.text(j, i, f"{corr.iloc[i, j]:.2f}", ha="center", va="center",
                    color="white" if abs(corr.iloc[i, j]) > 0.5 else "black", fontsize=8)
    ax.set_title("Heatmap de Correlação entre Variáveis Numéricas")
    fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
    salvar(fig, "12_heatmap_correlacao.png")

    print(f"{len(list(GRAPH_DIR.glob('*.png')))} gráficos salvos em {GRAPH_DIR}")
    print()


# ----------------------------------------------------------------------------
# ETAPA 9 — ARQUIVOS FINAIS
# ----------------------------------------------------------------------------
def salvar_planilhas(df_limpo_export: pd.DataFrame, df_analisado: pd.DataFrame,
                      geral: pd.Series, por_provedor: pd.DataFrame, por_bairro: pd.DataFrame,
                      por_periodo: pd.DataFrame, df_suspeitos: pd.DataFrame) -> None:
    print("=" * 70)
    print("ETAPA 9 — GERAÇÃO DOS ARQUIVOS FINAIS")
    print("=" * 70)

    caminho_limpo = CLEAN_DATA_FILE
    df_limpo_export.to_excel(caminho_limpo, index=False, sheet_name="dados_limpos")
    print(f"Salvo: {caminho_limpo}")

    caminho_analisado = ANALYSIS_FILE
    with pd.ExcelWriter(caminho_analisado, engine="openpyxl") as writer:
        df_analisado.to_excel(writer, index=False, sheet_name="dados_analisados")
        geral.to_frame(name="valor").to_excel(writer, sheet_name="estatisticas_gerais")
        por_provedor.to_excel(writer, index=False, sheet_name="por_provedor")
        por_bairro.to_excel(writer, index=False, sheet_name="por_bairro")
        por_periodo.to_excel(writer, index=False, sheet_name="por_periodo")
        if len(df_suspeitos):
            df_suspeitos.to_excel(writer, index=False, sheet_name="dados_suspeitos")
    print(f"Salvo: {caminho_analisado}")

    # formatação básica (fonte/largura de colunas) via openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    for caminho in (caminho_limpo, caminho_analisado):
        wb = load_workbook(caminho)
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.font = Font(name="Arial", bold=True)
            for col_cells in ws.columns:
                largura = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max(largura + 2, 10), 45)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = Font(name="Arial")
        wb.save(caminho)
    print("Formatação (fonte Arial, largura de colunas) aplicada.")
    print()


def _df_to_markdown(df: pd.DataFrame) -> str:
    """Converte DataFrame para tabela Markdown sem precisar do pacote tabulate."""
    cols = list(df.columns)
    rows = df.astype(str).values.tolist()
    widths = [max(len(c), max((len(r[i]) for r in rows), default=0)) for i, c in enumerate(cols)]
    sep = "| " + " | ".join("-" * w for w in widths) + " |"
    header = "| " + " | ".join(c.ljust(widths[i]) for i, c in enumerate(cols)) + " |"
    body = "\n".join(
        "| " + " | ".join(r[i].ljust(widths[i]) for i in range(len(cols))) + " |"
        for r in rows
    )
    return "\n".join([header, sep, body])


def gerar_relatorio(df: pd.DataFrame, geral: pd.Series, por_provedor: pd.DataFrame,
                     por_bairro: pd.DataFrame, por_periodo: pd.DataFrame,
                     df_suspeitos: pd.DataFrame, insights: dict, n_inicial: int) -> None:
    linhas = []
    a = linhas.append

    a("# Relatório Técnico — Qualidade da Internet em Caxias-MA\n")
    a(f"**Respostas recebidas (planilha original):** {n_inicial}  ")
    a(f"**Respostas válidas após limpeza:** {geral['qtd_respostas']:.0f}  ")
    a(f"**Bairros representados:** {geral['qtd_bairros']:.0f}  ")
    a(f"**Provedores avaliados:** {geral['qtd_provedores']:.0f}\n")

    a("## 1. Metodologia\n")
    a("Os dados foram coletados via formulário do Google Forms, exportados em `.xlsx` e "
      "processados com um pipeline em Python (pandas/numpy/matplotlib/openpyxl). O formulário "
      "continha duas rodadas de medição (download, upload, ping e perda de pacotes) por "
      "respondente; os valores finais utilizados na análise são a **média das duas medições**. "
      "Como as respostas de velocidade, ping e perda de pacotes foram coletadas em **faixas "
      "categóricas** (ex.: \"200 a 399 Mbps\", \"30 a 49 ms\"), cada faixa foi convertida em um "
      "valor numérico representativo (ponto médio da faixa; faixas abertas como \"Acima de 10%\" "
      "ou \"Menos de 10 Mbps\" foram tratadas com critério conservador — ver seção de limitações). "
      "Não havia, na exportação fornecida, colunas de *traceroute* ou de horário explícito da "
      "medição; o período do dia (manhã/tarde/noite) foi derivado do carimbo de data/hora de envio "
      "do formulário, usado como proxy do horário da medição.\n")

    a("## 2. Limpeza realizada\n")
    a("- Linhas totalmente vazias e duplicadas exatas foram removidas.")
    a("- A coluna de consentimento (irrelevante para análise, resposta única) foi descartada.")
    a("- Nomes de bairros e provedores foram padronizados (remoção de espaços extras, acentuação "
      "e maiúsculas/minúsculas inconsistentes; variações como \"Vila paraiso\", \"residencial vila "
      "paraíso\" foram unificadas em \"Vila Paraíso\").")
    a("- Tipos de conexão foram padronizados (ex.: variações de \"fibra óptica\" unificadas).")
    a("- Faixas de velocidade/ping/perda foram convertidas para valores numéricos únicos (Mbps, ms, %).")
    a("- Registros sem nenhuma métrica numérica válida (download, upload, ping e perda todos "
      "ausentes) foram removidos por impossibilitarem a análise.")
    a(f"- Total de **{n_inicial - geral['qtd_respostas']:.0f}** registro(s) removido(s) na limpeza, "
      f"restando **{geral['qtd_respostas']:.0f}** respostas válidas.\n")

    a("## 3. Dados suspeitos\n")
    if len(df_suspeitos):
        a(f"Foram identificados **{df_suspeitos['indice'].nunique()}** registro(s) com valores "
          f"suspeitos (mantidos na base, não excluídos automaticamente):\n")
        for motivo, qtd in df_suspeitos["motivo"].value_counts().items():
            a(f"- {motivo}: {qtd} ocorrência(s)")
        a("")
    else:
        a("Nenhum dado suspeito foi identificado nos critérios avaliados (valores negativos, "
          "perda fora de 0–100%, velocidades extremas, download muito acima do plano contratado).\n")

    a("## 4. Estatísticas gerais\n")
    a(f"- Quantidade de respostas válidas: **{geral['qtd_respostas']:.0f}**")
    a(f"- Quantidade de bairros: **{geral['qtd_bairros']:.0f}**")
    a(f"- Quantidade de provedores: **{geral['qtd_provedores']:.0f}**")
    a(f"- Download médio: **{geral['download_medio_mbps']:.1f} Mbps**")
    a(f"- Upload médio: **{geral['upload_medio_mbps']:.1f} Mbps**")
    a(f"- Ping médio: **{geral['ping_medio_ms']:.1f} ms**")
    a(f"- Perda de pacotes média: **{geral['perda_media_pct']:.2f}%**\n")

    a("## 5. Estatísticas por provedor (do melhor para o pior)\n")
    a(_df_to_markdown(por_provedor))
    a("")

    a("## 6. Estatísticas por bairro\n")
    a(_df_to_markdown(por_bairro))
    a("")

    a("## 7. Comparação por período do dia\n")
    a(_df_to_markdown(por_periodo))
    a("")
    if insights.get("degradacao_horario_pico") is not None:
        if insights["degradacao_horario_pico"]:
            a(f"Há indício de **degradação de ping no horário de pico** ({insights.get('periodo_maior_ping')}), "
              f"com diferença de {insights.get('diferenca_ping_periodos_ms')} ms em relação ao período de "
              f"melhor desempenho ({insights.get('periodo_menor_ping')}).\n")
        else:
            a(f"Não há indício relevante de degradação por horário de pico nesta amostra "
              f"(diferença de ping entre períodos de apenas {insights.get('diferenca_ping_periodos_ms')} ms).\n")

    a("## 8. Padrões identificados\n")
    a(f"- **Melhor provedor (maior score de classificação):** {insights.get('melhor_provedor', '—')}")
    a(f"- **Pior provedor:** {insights.get('pior_provedor', '—')}")
    a(f"- **Menor ping médio:** {insights.get('menor_ping_provedor', '—')}")
    a(f"- **Maior aproveitamento do plano contratado:** {insights.get('maior_aproveitamento_provedor', '—')}")
    if "maior_variacao_provedor" in insights:
        a(f"- **Maior variação de download (desvio padrão) entre respondentes:** {insights['maior_variacao_provedor']}")
    a(f"- **Bairro com melhor desempenho médio:** {insights.get('melhor_bairro', '—')}")
    a(f"- **Bairro com pior desempenho médio:** {insights.get('pior_bairro', '—')}")
    if insights.get("bairros_desempenho_muito_inferior"):
        a(f"- **Bairro(s) com desempenho muito inferior à média geral:** "
          f"{', '.join(insights['bairros_desempenho_muito_inferior'])}")
    else:
        a("- Nenhum bairro apresentou desempenho extremamente inferior aos demais (abaixo de 50% da média geral).")
    if "diferenca_significativa_bairros" in insights:
        diff_txt = "há diferença significativa" if insights["diferenca_significativa_bairros"] else \
            "não há diferença muito acentuada"
        a(f"- Considerando o coeficiente de variação do download médio entre bairros "
          f"({insights['coef_variacao_bairros']}), **{diff_txt}** de desempenho entre os bairros.")
    a("")

    a("## 9. Classificação geral das conexões\n")
    contagem_classif = df["classificacao"].value_counts()
    for classe, qtd in contagem_classif.items():
        pct = qtd / len(df) * 100
        a(f"- {classe}: {qtd} ({pct:.1f}%)")
    a("")
    a("**Critérios de classificação utilizados** (baseados no aproveitamento do plano contratado, "
      "ping e perda de pacotes):\n")
    a("- Ótima: aproveitamento ≥ 90% do plano, ping ≤ 20 ms, perda = 0%")
    a("- Boa: aproveitamento ≥ 75%, ping ≤ 40 ms, perda ≤ 1%")
    a("- Regular: aproveitamento ≥ 50%, ping ≤ 70 ms, perda ≤ 2%")
    a("- Ruim: aproveitamento ≥ 30%, ping ≤ 100 ms, perda ≤ 5%")
    a("- Péssima: abaixo desses valores")
    a("- Não classificado: faltam dados (plano contratado, ping ou perda desconhecidos)\n")

    a("## 10. Gráficos gerados\n")
    a("Todos os gráficos estão disponíveis na pasta `graficos/` em alta resolução (300 DPI):\n")
    for f in sorted(GRAPH_DIR.glob("*.png")):
        a(f"- `{f.name}`")
    a("")

    a("## 11. Conclusões\n")
    a(f"A amostra coletada ({geral['qtd_respostas']:.0f} respostas válidas, {geral['qtd_bairros']:.0f} "
      f"bairros e {geral['qtd_provedores']:.0f} provedores) indica que a média de download na cidade "
      f"fica em torno de {geral['download_medio_mbps']:.1f} Mbps, com ping médio de "
      f"{geral['ping_medio_ms']:.1f} ms. O provedor **{insights.get('melhor_provedor', '—')}** se destacou "
      f"com o melhor desempenho médio combinado, enquanto **{insights.get('pior_provedor', '—')}** "
      f"apresentou o pior resultado entre os avaliados. ")
    if insights.get("bairros_desempenho_muito_inferior"):
        a(f"Os bairros {', '.join(insights['bairros_desempenho_muito_inferior'])} merecem atenção por "
          f"apresentarem desempenho médio bem abaixo da média da cidade.")
    a("")

    a("## 12. Limitações dos dados\n")
    a("- **Amostra pequena e não probabilística**: poucos respondentes por provedor/bairro, o que "
      "limita a significância estatística das comparações (alguns provedores têm apenas 1 ou 2 "
      "respondentes).")
    a("- **Dados em faixas, não valores exatos**: download, upload, ping e perda de pacotes foram "
      "coletados como faixas categóricas no formulário, não como medições exatas; os valores usados "
      "na análise são pontos médios estimados dessas faixas, o que introduz imprecisão.")
    a("- **Faixas abertas**: respostas como \"Acima de 10%\" ou \"Menos de 10 Mbps\" não têm limite "
      "definido; foram usadas estimativas conservadoras (valor mínimo citado, ou metade dele), que "
      "podem não refletir o valor real.")
    a("- **Autodeclaração**: os dados são autorrelatados pelos usuários, sem verificação técnica "
      "independente (ex.: speedtest padronizado), sujeitos a erro de leitura/interpretação.")
    a("- **Sem coluna de horário real da medição**: o período do dia foi inferido a partir do horário "
      "de envio do formulário, que pode não coincidir com o horário em que o teste de internet foi "
      "efetivamente realizado.")
    a("- **Sem dados de traceroute**: a planilha fornecida não continha informações de traceroute, "
      "impossibilitando essa análise.")
    a("- **Possível viés de seleção**: usuários insatisfeitos podem ter maior propensão a responder "
      "pesquisas sobre qualidade de internet, o que pode enviesar os resultados para baixo.")

    REPORT_FILE.write_text("\n".join(linhas), encoding="utf-8")
    print(f"Relatório salvo em: {REPORT_FILE}")


# ----------------------------------------------------------------------------
# MAIN
# ----------------------------------------------------------------------------
def main():
    df_raw = pd.read_excel(INPUT_FILE)
    n_inicial = len(df_raw)

    inspecionar(df_raw)
    df_limpo = limpar_dados(df_raw)
    df_suspeitos = detectar_suspeitos(df_limpo)
    df_comcolunas = criar_colunas(df_limpo)
    df_analisado = classificar_conexoes(df_comcolunas)

    print("=" * 70)
    print("ETAPA 6 — ESTATÍSTICAS")
    print("=" * 70)
    geral = estatisticas_gerais(df_analisado)
    por_provedor = estatisticas_por_provedor(df_analisado)
    por_bairro = estatisticas_por_bairro(df_analisado)
    por_periodo = estatisticas_por_periodo(df_analisado)
    print("Estatísticas gerais:\n", geral, "\n")
    print("Por provedor:\n", por_provedor, "\n")
    print("Por bairro:\n", por_bairro, "\n")
    print("Por período:\n", por_periodo, "\n")

    print("=" * 70)
    print("ETAPA 7 — PADRÕES")
    print("=" * 70)
    insights = descobrir_padroes(df_analisado, por_provedor, por_bairro, por_periodo)
    for k, v in insights.items():
        print(f"  {k}: {v}")
    print()

    gerar_graficos(df_analisado, por_provedor, por_bairro)

    salvar_planilhas(df_limpo, df_analisado, geral, por_provedor, por_bairro, por_periodo, df_suspeitos)

    gerar_relatorio(df_analisado, geral, por_provedor, por_bairro, por_periodo,
                     df_suspeitos, insights, n_inicial)

    print("\nPIPELINE CONCLUÍDO COM SUCESSO.\n")
    return df_analisado, geral, por_provedor, por_bairro, por_periodo, insights


if __name__ == "__main__":
    main()
